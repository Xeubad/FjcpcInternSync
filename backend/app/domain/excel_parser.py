import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def generate_excel_template():
    """生成Excel模板文件"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    example_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    example_font = Font(italic=True, color="FF666666", size=10)

    for report_type, sheet_name in [
        ("day", "日报"),
        ("week", "周报"),
        ("month", "月报"),
    ]:
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "是否上传(是/否)",
            "日期(YYYY-MM-DD)",
            "实习工作具体情况及实习任务完成情况",
            "主要收获及工作成绩",
            "工作中的问题及需要老师的指导帮助",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        example_data = {
            "日报": [
                "是",
                "2025-01-13",
                "在XXX公司进行实习，主要负责XXX工作，完成了XXX任务",
                "通过实习，我学到了XXX技能，提高了XXX能力",
                "暂无问题",
            ],
            "周报": [
                "是",
                "2025-01-10",
                "本周在XXX部门实习，主要参与XXX项目，完成XXX工作",
                "本周收获：1.XXX 2.XXX 3.XXX",
                "希望老师指导XXX",
            ],
            "月报": [
                "是",
                "2025-01-31",
                "本月在XXX公司实习，主要负责XXX工作，本月完成XXX任务",
                "本月主要收获：1.XXX 2.XXX 3.XXX",
                "无",
            ],
        }

        for col, value in enumerate(example_data[sheet_name], 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        col_count = 5
        for col in range(1, col_count + 1):
            ws.column_dimensions[get_column_letter(col)].width = 35
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


TEST_KEYWORDS = [
    "示例",
    "例子",
    "测试",
    "测试内容",
    "请填写",
    "请输入",
    "XXX",
    "xxx",
    "这里填写",
    "参照此格式",
    "示例内容",
    "测试数据",
    "demo",
    "test",
]


def check_test_content(ws, sheet_name):
    """检查第2行是否包含示例数据（通过颜色检测）"""
    cell = ws.cell(row=2, column=2)
    fill = cell.fill

    if fill and fill.start_color and fill.start_color.rgb:
        color = fill.start_color.rgb
        if color == "FFF2CC" or color == "FFFFF2CC":
            return True, f'工作表"{sheet_name}"第2行示例数据未删除，请删除第2行后再上传'

    first_row = ws.cell(row=1, column=2).value
    if not first_row or "日期" not in str(first_row):
        return True, f'工作表"{sheet_name}"表头被删除，请使用正确模板'

    return False, None


def parse_excel_sheet(ws, sheet_name):
    """解析Excel工作表"""
    reports = []

    type_map = {"日报": "day", "周报": "week", "月报": "month"}
    report_type = type_map.get(sheet_name, sheet_name)

    if ws.max_row < 2:
        logger.info(f'工作表 "{sheet_name}" 无数据，跳过')
        return reports, 0

    from datetime import datetime

    logger.info(f'开始解析工作表 "{sheet_name}"，最大行数: {ws.max_row}')

    skipped_no = 0
    for row_idx in range(2, ws.max_row + 1):
        upload_flag = ws.cell(row=row_idx, column=1).value
        report_date = ws.cell(row=row_idx, column=2).value
        work_content = ws.cell(row=row_idx, column=3).value
        achievement = ws.cell(row=row_idx, column=4).value
        problem = ws.cell(row=row_idx, column=5).value

        logger.debug(
            f"行{row_idx}: upload_flag={repr(upload_flag)}, date={repr(report_date)}, has_work={bool(work_content)}"
        )

        if not report_date and not work_content and not achievement and not problem:
            continue

        has_content = bool(work_content or achievement or problem)

        if has_content:
            flag_value = str(upload_flag).strip() if upload_flag is not None else ""
            flag_lower = flag_value.lower()

            valid_yes_flags = ["是", "yes", "true", "1", "y", "t"]
            valid_no_flags = ["否", "no", "false", "0", "n", "f"]

            if flag_lower in valid_yes_flags:
                pass
            elif flag_lower in valid_no_flags:
                skipped_no += 1
                continue
            else:
                if upload_flag is None:
                    raise ValueError(
                        f'工作表"{sheet_name}"第{row_idx}行"是否上传"选项为空，请填写"是"表示上传，填写"否"表示不上传。'
                    )
                else:
                    raise ValueError(
                        f'工作表"{sheet_name}"第{row_idx}行"是否上传"选项无效：检测到值{repr(upload_flag)}，请填写"是"（或"1"、"true"、"yes"）表示上传，填写"否"（或"0"、"false"、"no"）表示不上传。'
                    )

            if report_date is None:
                raise ValueError(
                    f'工作表"{sheet_name}"第{row_idx}行日期为空，请填写日期（格式：YYYY-MM-DD）'
                )

            if isinstance(report_date, datetime):
                date_obj = report_date
            elif isinstance(report_date, str):
                try:
                    date_obj = datetime.strptime(report_date, "%Y-%m-%d")
                except ValueError:
                    raise ValueError(
                        f'工作表"{sheet_name}"第{row_idx}行日期格式错误 "{report_date}"，应为YYYY-MM-DD格式'
                    )
            else:
                raise ValueError(
                    f'工作表"{sheet_name}"第{row_idx}行日期格式错误，应为YYYY-MM-DD格式'
                )

            work_text = str(work_content or "")
            achieve_text = str(achievement or "")
            problem_text = str(problem or "")

            total_length = len(work_text) + len(achieve_text) + len(problem_text)

            if total_length < 200:
                raise ValueError(
                    f'工作表"{sheet_name}"第{row_idx}行内容不足200字（当前{total_length}字），请补充内容后再上传'
                )

            reports.append(
                {
                    "type": report_type,
                    "date": date_obj,
                    "work": work_content or "",
                    "achievement": achievement or "",
                    "problem": problem or "",
                    "word_count": total_length,
                }
            )

    if skipped_no > 0:
        logger.info(
            f'工作表 "{sheet_name}" 解析到 {len(reports)} 条记录（另有 {skipped_no} 行填写"否"已跳过）'
        )
    else:
        logger.info(f'工作表 "{sheet_name}" 解析到 {len(reports)} 条记录')
    return reports, skipped_no


def parse_excel_file(file_content):
    """解析Excel文件，返回日报/周报/月报列表"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))

        result = {"day": [], "week": [], "month": []}

        sheet_map = {
            "日报": "day",
            "周报": "week",
            "月报": "month",
            "day": "day",
            "week": "week",
            "month": "month",
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_key = sheet_map.get(sheet_name, sheet_name)

            if sheet_key in result:
                has_test, test_msg = check_test_content(ws, sheet_name)
                if has_test:
                    raise ValueError(test_msg)

                reports, skipped_no = parse_excel_sheet(ws, sheet_name)
                result[sheet_key] = reports
                result[f"{sheet_key}_skipped"] = skipped_no

        for report_type, reports_list in [
            ("day", result["day"]),
            ("week", result["week"]),
            ("month", result["month"]),
        ]:
            if not reports_list:
                continue
            seen_dates = set()
            for i, report in enumerate(reports_list):
                date_val = report["date"]
                if hasattr(date_val, "strftime"):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                if date_str in seen_dates:
                    type_name = {"day": "日报", "week": "周报", "month": "月报"}[
                        report_type
                    ]
                    raise ValueError(
                        f'工作表"{type_name}"中存在重复日期 "{date_str}"，请检查并修正'
                    )
                seen_dates.add(date_str)

        return result

    except ValueError:
        raise
    except Exception as e:
        logger.error(f"解析Excel文件失败: {e}")
        raise
